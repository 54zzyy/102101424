# 1确定数据来源
# 2通过开发者工具进行抓包分析、

import requests  # 第三方
import re  # 正则表达式
import openpyxl
import heapq
import cProfile
import pstats
def run_code():
    cnts={}
    for page in range(1,11):
        if page==1:
            url=f'https://search.bilibili.com/video?vt=12970522&keyword=日本核污染水排海'
        else:
            url=f'https://search.bilibili.com/video?vt=12970522&keyword=日本核污染水排海&page={page}'
    # headers 请求头，进行伪装
        headers = {
            'cookie':'buvid_fp=51715715b2a25511c04e04900c67ac89; CURRENT_FNVAL=4048; rpdid=|(JlklRl)~ll0J\'uY~||YlYJ); buvid3=C2010A62-0C1A-F8CB-068D-84F6663B8B0F56336infoc; b_nut=1690364956; i-wanna-go-back=-1; b_ut=7; _uuid=D268E9D2-E7B9-41D9-109FF-10BE4D585D1A959926infoc; FEED_LIVE_VERSION=V8; home_feed_column=5; browser_resolution=1494-798; header_theme_version=CLOSE; DedeUserID=1979916649; DedeUserID__ckMd5=94614b2a9e4d26e6; bp_video_offset_1979916649=833215220115570691; buvid4=3B3C49DA-5B87-6BEF-BB22-F6DC43F29BA239435-022060222-SK3hbof5R8luR%2Blb5tSuihcv1nabILCmxo5Tq3B%2FuIfkHyGhIuUb3w%3D%3D; SESSDATA=ffe603c5%2C1710073072%2C89b80%2A91CjC3xvLDXGJ4ZnljmhSTfTL3NzOvLm3hERvD_a4sOM2er_wcNSw8-p5UJhi78qzID2YSVkFLN2JDa0NFam5nTzFqeFVMaXk1X0lvWWh0aUFrbC00S3BmZmliVmhvTFhOLWpjdy1aRVhPWjhWV0FESkJ0YkRpRVljOHRBc1g0YVJfTmo2SWJjeTJBIIEC; bili_jct=9461aa8599e76ec432014a05835638d9; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2OTQ3ODIxOTYsImlhdCI6MTY5NDUyMjk5NiwicGx0IjotMX0.47RVAAuRvp2vipJCLGZP6B6Ya7Oz96YvAwmpRrwawPM; bili_ticket_expires=1694782196; PVID=1; b_lsid=88588489_18A8ECDC0CA; sid=7vjil106',
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 SLBrowser/8.0.1.5162 SLBChan/103'
        }
        response = requests.get(url=url, headers=headers)
        response.encoding = response.apparent_encoding
        #print(response.text)
        text_list = re.findall(',aid:.*?bvid:"(.*?)",title:.*?keyword', response.text)
        for i in text_list:
            video_url ='http://www.ibilibili.com/video/'+ i
            #print(video_url)
            resp = requests.get(url=video_url)
            resp.encoding = resp.apparent_encoding
            #print(resp.text)
            dm_url = re.findall('<a href="(.*?)"  class="btn btn-default" target="_blank">弹幕</a>', resp.text)
            '''for index in dm_url:
                print(index)'''
            for index in dm_url:
                resp_new = requests.get(url=index)
                resp_new.encoding = resp.apparent_encoding
                data_list = re.findall('<d p=".*?">(.*?)</d>', resp_new.text)
                for i in data_list:
                    with open('弹幕.txt', mode='a', encoding='utf-8') as f:
                        f.write(i)
                        f.write('\n')
                        if i in cnts:
                            cnts[i] += 1
                        else:
                            cnts[i] = 1
                        print(i)
        def write_lines_excel(arr):
            work_book = openpyxl.Workbook()
            sheet = work_book.create_sheet('arrage')
            sheet.cell(1, 1, 'top20的弹幕内容')
            sheet.cell(1, 2, '弹幕出现次数')
            for index, row in enumerate(arr):
                for co in range(len(row)):
                    sheet.cell(index + 2, co + 1, row[co])
            work_book.save('top20弹幕.xlsx')
    dic = {}
    dic.update({k: cnts[k] for k in heapq.nlargest(20, cnts, key=cnts.get)})
    arr=[]
    for key,value in dic.items():
        arr.append((key,value))
    #将top20的弹幕生成表格
    write_lines_excel(arr)

# 创建性能分析器对象
profiler = cProfile.Profile()

# 执行代码并进行性能分析
profiler.runctx('run_code()', globals(), locals())

# 获取函数执行时间统计信息
stats = profiler.getstats()

# 输出与当前代码相关的函数的执行时间
for stat in stats:
    code = stat.code
    if code.co_filename == __file__:
        print(f"函数 {code.co_name} 的执行时间为: {stat.totaltime:.4f} 秒")